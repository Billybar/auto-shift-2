import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ============================================================================
# 1. Style Setup
# ============================================================================
def _setup_styles():
    """
    Creates and returns a dictionary of all OpenPyXL styles used in the workbook.
    """
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    return {
        'header_font': Font(bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'sub_header_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'stats_header_fill': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'orange_header_fill': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        'blue_header_fill': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
        'purple_header_fill': PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
        'red_header_fill': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        # New for penalties
        'grey_fill': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),

        # Alerts / Conditional Formatting
        'alert_red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'alert_green': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'alert_orange': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),

        'border': thin_border,
        'center': center_align,
        'right': Alignment(horizontal='right', vertical='center'),
        'bold': Font(bold=True)
    }


# ============================================================================
# 2. Logic Helpers
# ============================================================================
def _get_shift_counts(solver, shift_vars, employees, num_days, num_shifts):
    """
    Calculates the raw number of shifts assigned to each employee by the solver.
    """
    counts = {i: {s: 0 for s in range(num_shifts)} for i in range(len(employees))}
    for e_idx in range(len(employees)):
        for d in range(num_days):
            for s in range(num_shifts):
                if solver.Value(shift_vars[(e_idx, d, s)]):
                    counts[e_idx][s] += 1
    return counts


def _get_employee_rank(employees, emp_idx):
    """Helper to sort employees by rank: Supervisor > Controller > Guard"""
    r = employees[emp_idx].get('role', 'guard')
    if r == 'supervisor': return 3
    if r == 'controller': return 2
    return 1


def _calculate_penalty_counts(solver, shift_vars, employees, num_days, num_shifts):
    """
    Scans the solution to count specific violations per employee.
    Returns: dict {emp_idx: {'chain_3': int, 'rest_gap': int, 'consecutive_nights': int}}
    """
    penalties = {i: {'chain_3': 0, 'rest_gap': 0, 'consecutive_nights': 0} for i in range(len(employees))}

    # Helper lambda to check if working specific shift
    is_working = lambda e, d, s: solver.Value(shift_vars[(e, d, s)]) == 1

    # Helper: Check if working Morning (0) OR Reinforcement (3)
    def is_morn_reinf(e, d):
        if d >= num_days: return False
        # Assuming Shift 0 is Morning and Shift 3 is Reinforcement (Morning overlap)
        return is_working(e, d, 0) or is_working(e, d, 3)

    for e in range(len(employees)):

        # 1. Count Consecutive Nights (3 in a row)
        for d in range(num_days - 2):
            if is_working(e, d, 2) and is_working(e, d + 1, 2) and is_working(e, d + 2, 2):
                penalties[e]['consecutive_nights'] += 1

        # 2. Count Rest Gaps (8-8 Patterns)
        # We explicitly check for the "Bad Combinations" of shifts that result in 8h rest only.
        # Combinations:
        # A. Same Day: Morning/Reinf -> Night
        # B. Next Day: Noon -> Morning/Reinf
        # C. Next Day: Night -> Noon

        for d in range(num_days):
            # A. Same Day: Start Early (0/3) -> End Late (2)
            if is_morn_reinf(e, d) and is_working(e, d, 2):
                penalties[e]['rest_gap'] += 1

            # Check Transitions to Next Day
            if d < num_days - 1:
                # B. Noon (1) -> Next Morning/Reinf (0/3)
                if is_working(e, d, 1) and is_morn_reinf(e, d + 1):
                    penalties[e]['rest_gap'] += 1

                # C. Night (2) -> Next Noon (1)
                if is_working(e, d, 2) and is_working(e, d + 1, 1):
                    penalties[e]['rest_gap'] += 1

        # 3. Count Chain 3 (The 8-8-8 Patterns: A, B, C)
        for d in range(num_days - 1):
            # Pattern A: (Morn/Reinf D) -> (Night D) -> (Noon D+1)
            if is_morn_reinf(e, d) and is_working(e, d, 2) and is_working(e, d + 1, 1):
                penalties[e]['chain_3'] += 1

            # Pattern B: (Noon D) -> (Morn/Reinf D+1) -> (Night D+1)
            if is_working(e, d, 1) and is_morn_reinf(e, d + 1) and is_working(e, d + 1, 2):
                penalties[e]['chain_3'] += 1

            # Pattern C: (Night D) -> (Noon D+1) -> (Morn/Reinf D+2)
            if d < num_days - 2:
                if is_working(e, d, 2) and is_working(e, d + 1, 1) and is_morn_reinf(e, d + 2):
                    penalties[e]['chain_3'] += 1

    return penalties


# ============================================================================
# 3. Sheet Generators
# ============================================================================
def _create_schedule_sheet(wb, styles, solver, shift_vars, employees, num_days, colors):
    """Generates the main 'Schedule' sheet."""
    ws = wb.active
    ws.title = "Schedule"
    ws.sheet_view.rightToLeft = True

    # --- Headers ---
    ws.cell(row=1, column=1).value = "זמן"
    ws.cell(row=1, column=2).value = "תפקיד"
    days_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]

    for i, day in enumerate(days_names):
        c = ws.cell(row=1, column=i + 3)
        c.value = day
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = styles['center']

    # Define the Table Layout
    layout = [
        # --- Building 15 ---
        {"title": "בניין 15", "is_header": True},
        {"time": "בוקר", "role": "קבלה", "shift": 0, "role_needed": "guard"},
        {"time": "", "role": "סייר", "shift": 0, "role_needed": "guard"},
        {"time": "", "role": "אחמ\"ש", "shift": 3, "role_needed": "supervisor"},
        {"is_spacer": True},
        {"time": "צהריים", "role": "קבלה", "shift": 1, "role_needed": "guard"},
        {"time": "", "role": "סייר", "shift": 1, "role_needed": "guard"},
        {"is_spacer": True},
        {"time": "לילה", "role": "קבלה", "shift": 2, "role_needed": "guard"},
        {"time": "", "role": "סייר", "shift": 2, "role_needed": "guard"},

        # --- Building 18 ---
        {"title": "בניין 18", "is_header": True},

        # Morning
        {"time": "בוקר", "role": "קבלה", "shift": 0, "role_needed": "guard"},
        {"time": "", "role": "בקרה", "shift": 0, "role_needed": "controller"},
        {"time": "", "role": "אחמ\"ש", "shift": 0, "role_needed": "supervisor"},
        # New row added for the 4th guard (Patrol/Sayer)
        {"time": "", "role": "סייר", "shift": 0, "role_needed": "guard"},

        # Reinforcement / Special (07-18)
        {"time": "", "role": "סייר (7-18)", "shift": 3, "role_needed": "guard"},
        {"time": "", "role": "מאבטח 1 (7-18)", "shift": 3, "role_needed": "guard"},
        {"time": "", "role": "מאבטח 2 (7-18)", "shift": 3, "role_needed": "guard"},
        {"is_spacer": True},

        # Noon
        {"time": "צהריים", "role": "קבלה", "shift": 1, "role_needed": "guard"},
        {"time": "", "role": "בקרה", "shift": 1, "role_needed": "controller"},
        {"time": "", "role": "אחמ\"ש", "shift": 1, "role_needed": "supervisor"},
        # New row added for the 4th guard (Patrol/Sayer)
        {"time": "", "role": "סייר", "shift": 1, "role_needed": "guard"},

        {"is_spacer": True},

        # Night
        {"time": "לילה", "role": "קבלה", "shift": 2, "role_needed": "guard"},
        {"time": "", "role": "בקרה", "shift": 2, "role_needed": "controller"},
        {"time": "", "role": "סייר", "shift": 2, "role_needed": "guard"},
    ]

    current_row = 2
    used_assignments = set()
    role_fill_counts = {i: {'supervisor': 0, 'controller': 0, 'guard': 0} for i in range(len(employees))}

    for row_def in layout:
        if row_def.get("is_header"):
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            c = ws.cell(row=current_row, column=1)
            c.value = row_def["title"]
            c.font = Font(bold=True, size=12)
            c.fill = styles['sub_header_fill']
            c.alignment = styles['center']
            current_row += 1
            continue

        if row_def.get("is_spacer"):
            current_row += 1
            continue

        time_c = ws.cell(row=current_row, column=1)
        time_c.value = row_def["time"]
        time_c.border = styles['border']
        time_c.alignment = styles['center']
        if row_def["time"]: time_c.font = styles['bold']

        role_c = ws.cell(row=current_row, column=2)
        role_c.value = row_def["role"]
        role_c.border = styles['border']
        role_c.alignment = styles['right']

        shift_idx = row_def["shift"]
        role_needed = row_def["role_needed"]

        for d in range(num_days):
            cell = ws.cell(row=current_row, column=d + 3)
            cell.border = styles['border']
            cell.alignment = styles['center']

            is_weekend = (d >= 5)
            skip = False
            if is_weekend and (shift_idx == 3 or "אחמ\"ש" in row_def["role"]):
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                skip = True

            if not skip:
                candidates = [e for e in range(len(employees)) if solver.Value(shift_vars[(e, d, shift_idx)])]

                if role_needed == 'guard':
                    candidates.sort(key=lambda x: _get_employee_rank(employees, x))
                elif role_needed == 'controller':
                    candidates.sort(key=lambda x: (0 if _get_employee_rank(employees, x) == 2 else 1))

                assigned_emp = None
                for cand_idx in candidates:
                    if (d, cand_idx) in used_assignments: continue

                    emp_role = employees[cand_idx].get('role', 'guard')
                    is_match = False
                    if role_needed == 'guard':
                        is_match = True
                    elif role_needed == 'controller' and emp_role in ['controller', 'supervisor']:
                        is_match = True
                    elif role_needed == 'supervisor' and emp_role == 'supervisor':
                        is_match = True

                    if is_match:
                        assigned_emp = cand_idx
                        used_assignments.add((d, cand_idx))
                        role_fill_counts[assigned_emp][role_needed] += 1
                        break

                if assigned_emp is not None:
                    name = employees[assigned_emp]['name']
                    c_code = colors[assigned_emp] if assigned_emp < len(colors) else "FFFFFF"
                    cell.value = name
                    cell.fill = PatternFill(start_color=c_code, end_color=c_code, fill_type="solid")

        current_row += 1

    return role_fill_counts


def _create_stats_sheet(wb, styles, employees, shift_counts, role_fill_counts):
    """Generates the 'Statistics' sheet."""
    ws = wb.create_sheet("סטטיסטיקות")
    ws.sheet_view.rightToLeft = True

    # --- Table 1: General Shift Counts ---
    row = 2
    ws.cell(row=row, column=1).value = "סיכום משמרות לעובד (כללי)"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1

    headers = ["שם העובד", "בוקר", "צהריים", "לילה", "תגבור", 'סה"כ כללי', 'יעד']
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=idx)
        c.value = h
        c.font = styles['header_font']
        c.fill = styles['stats_header_fill']
        c.alignment = styles['center']
        c.border = styles['border']
    row += 1

    totals = [0, 0, 0, 0, 0, 0]

    for e_idx, emp in enumerate(employees):
        counts = shift_counts[e_idx]
        total = sum(counts.values())
        target = emp.get('target_shifts', 0)

        totals[0] += counts[0];
        totals[1] += counts[1];
        totals[2] += counts[2];
        totals[3] += counts[3]
        totals[4] += total;
        totals[5] += target

        row_data = [emp['name'], counts[0], counts[1], counts[2], counts[3], total, target]

        fill = None
        diff = target - total
        if diff < 0:
            fill = styles['alert_green']
        elif diff > 2:
            fill = styles['alert_red']
        if counts[2] > 2: fill = styles['alert_orange']

        for idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=idx)
            c.value = val
            c.border = styles['border']
            c.alignment = styles['center']
            if fill: c.fill = fill
        row += 1

    summary_vals = ["TOTAL"] + totals
    for idx, val in enumerate(summary_vals, 1):
        c = ws.cell(row=row, column=idx)
        c.value = val
        c.font = styles['bold']
        c.fill = styles['grey_fill']
        c.border = styles['border']
        c.alignment = styles['center']
    row += 3

    # --- Table 2: Unified Special Shifts (Supervisor & Controller) ---
    _create_unified_stats_table(ws, row, styles, employees, role_fill_counts)


def _create_unified_stats_table(ws, start_row, styles, employees, role_fill_counts):
    ws.cell(row=start_row, column=1).value = "סיכום משמרות בכירים (אחמ\"ש + בקרה)"
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
    start_row += 1

    headers = ["שם העובד", "תפקיד", "משמרות אחמ\"ש", "משמרות בקרה", "סה\"כ בכירים", "יעד"]
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=idx)
        c.value = h
        c.font = styles['header_font']
        c.fill = styles['purple_header_fill']
        c.border = styles['border']
        c.alignment = styles['center']
    start_row += 1

    for e_idx, emp in enumerate(employees):
        role = emp.get('role', 'guard')
        if role not in ['supervisor', 'controller']:
            continue

        role_display = "אחמ\"ש" if role == 'supervisor' else "בקר"
        sup_count = role_fill_counts[e_idx]['supervisor']
        ctrl_count = role_fill_counts[e_idx]['controller']
        total_special = sup_count + ctrl_count
        target = emp.get('target_shifts', 0)

        row_data = [emp['name'], role_display, sup_count, ctrl_count, total_special, target]

        fill = None
        if total_special == target:
            fill = styles['alert_green']
        elif abs(total_special - target) >= 2:
            fill = styles['alert_red']

        for idx, val in enumerate(row_data, 1):
            c = ws.cell(row=start_row, column=idx)
            c.value = val
            c.border = styles['border']
            c.alignment = styles['center']
            if idx in [5, 6] and fill: c.fill = fill
        start_row += 1


def _create_penalty_sheet(wb, styles, employees, penalty_data):
    """
    Creates a dedicated sheet for penalties (Rest Gap, Chain 3, Consecutive Nights).
    Only lists employees who have at least one penalty.
    """
    ws = wb.create_sheet("דוח חריגות ושחיקה")
    ws.sheet_view.rightToLeft = True

    row = 2
    ws.cell(row=row, column=1).value = "דוח חריגות ושחיקה"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1

    headers = ["שם העובד", "רצף שחיקה כבד (Chain 3)", "מרווח מנוחה קצר (Rest Gap)", "3 לילות ברצף"]

    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=idx)
        c.value = h
        c.font = styles['header_font']
        c.fill = styles['red_header_fill']
        c.alignment = styles['center']
        c.border = styles['border']
        # Set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 25
    row += 1

    has_content = False

    for e_idx, emp in enumerate(employees):
        p = penalty_data[e_idx]
        total_p = p['chain_3'] + p['rest_gap'] + p['consecutive_nights']

        if total_p > 0:
            has_content = True
            row_data = [emp['name'], p['chain_3'], p['rest_gap'], p['consecutive_nights']]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=c_idx)
                cell.value = val
                cell.border = styles['border']
                cell.alignment = styles['center']

                # Highlight specific cells with issues
                if c_idx > 1 and isinstance(val, int) and val > 0:
                    cell.fill = styles['alert_red']
                    cell.font = styles['bold']

            row += 1

    if not has_content:
        ws.cell(row=row, column=1).value = "לא נמצאו חריגות."


# ============================================================================
# 4. Main Entry Point
# ============================================================================
def create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, colors):
    # 1. Init Workbook & Styles
    wb = openpyxl.Workbook()
    styles = _setup_styles()

    # 2. Create Schedule Sheet & Get Role Stats
    role_fill_counts = _create_schedule_sheet(wb, styles, solver, shift_vars, employees, num_days, colors)

    # 3. Calculate Shift Counts (Raw data from solver)
    shift_counts = _get_shift_counts(solver, shift_vars, employees, num_days, num_shifts)

    # 4. Calculate Penalties (NEW)
    penalty_counts = _calculate_penalty_counts(solver, shift_vars, employees, num_days, num_shifts)

    # 5. Create Statistics Sheet
    _create_stats_sheet(wb, styles, employees, shift_counts, role_fill_counts)

    # 6. Create Penalty Sheet (NEW)
    _create_penalty_sheet(wb, styles, employees, penalty_counts)

    # 7. Save
    output_filename = "shift_schedule_output/shift_schedule_colored.xlsx"
    wb.save(output_filename)
    print(f"Excel file created successfully: {output_filename}")